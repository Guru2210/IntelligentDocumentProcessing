"""
Export utilities — JSON, CSV, Excel export for extraction results.
"""
import io
import json
from typing import Dict, Any, List


def export_to_json(result: Dict[str, Any]) -> bytes:
    return json.dumps(result, indent=2, default=str).encode("utf-8")


def export_to_csv(result: Dict[str, Any]) -> bytes:
    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    fields = result.get("fields", {})

    # Flat fields header row
    flat_fields = {k: v for k, v in fields.items() if v.get("type") != "array"}
    table_fields = {k: v for k, v in fields.items() if v.get("type") == "array"}

    if flat_fields:
        writer.writerow(["Field", "Value", "Confidence"])
        for field_name, field_data in flat_fields.items():
            value = (
                field_data.get("valueString")
                or field_data.get("valueNumber")
                or field_data.get("valueDate")
                or field_data.get("valueInteger")
                or ""
            )
            writer.writerow([field_name, value, field_data.get("confidence", "")])

    for table_name, table_data in table_fields.items():
        writer.writerow([])
        writer.writerow([f"TABLE: {table_name}"])
        rows = table_data.get("valueArray", [])
        if not rows:
            continue
        # Collect all column names
        all_cols = []
        for row in rows:
            for col in row.get("valueObject", {}).keys():
                if col not in all_cols:
                    all_cols.append(col)
        writer.writerow(all_cols + ["row_confidence"])
        for row in rows:
            obj = row.get("valueObject", {})
            values = [
                obj.get(col, {}).get("valueString") or obj.get(col, {}).get("valueNumber") or ""
                for col in all_cols
            ]
            row_conf = round(
                sum(obj.get(col, {}).get("confidence", 0) for col in all_cols) / max(len(all_cols), 1), 3
            )
            writer.writerow(values + [row_conf])

    return output.getvalue().encode("utf-8")


def export_to_excel(result: Dict[str, Any]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    fields = result.get("fields", {})
    flat_fields = {k: v for k, v in fields.items() if v.get("type") != "array"}
    table_fields = {k: v for k, v in fields.items() if v.get("type") == "array"}

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")

    headers = ["Field", "Value", "Type", "Confidence"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for field_name, field_data in flat_fields.items():
        value = (
            field_data.get("valueString")
            or field_data.get("valueNumber")
            or field_data.get("valueDate")
            or field_data.get("valueInteger")
            or ""
        )
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=2, value=str(value))
        ws.cell(row=row, column=3, value=field_data.get("type", "string"))
        conf = field_data.get("confidence", 0)
        conf_cell = ws.cell(row=row, column=4, value=conf)
        if isinstance(conf, (int, float)):
            if conf >= 0.9:
                conf_cell.font = Font(color="166534")
            elif conf >= 0.7:
                conf_cell.font = Font(color="92400E")
            else:
                conf_cell.font = Font(color="991B1B")
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12

    # Table sheets
    for table_name, table_data in table_fields.items():
        sheet_name = table_name[:31]  # Excel max sheet name length
        ws_t = wb.create_sheet(title=sheet_name)
        rows = table_data.get("valueArray", [])
        if not rows:
            continue

        all_cols = []
        for r in rows:
            for col in r.get("valueObject", {}).keys():
                if col not in all_cols:
                    all_cols.append(col)

        for col_idx, col_name in enumerate(all_cols, 1):
            cell = ws_t.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(rows, 2):
            obj = row_data.get("valueObject", {})
            for col_idx, col_name in enumerate(all_cols, 1):
                cell_data = obj.get(col_name, {})
                value = (
                    cell_data.get("valueString")
                    or cell_data.get("valueNumber")
                    or cell_data.get("valueDate")
                    or ""
                )
                ws_t.cell(row=row_idx, column=col_idx, value=value)

        for col_idx in range(1, len(all_cols) + 1):
            ws_t.column_dimensions[get_column_letter(col_idx)].width = 20

    # Overall info sheet
    ws_info = wb.create_sheet(title="Extraction Info")
    ws_info["A1"] = "Model Type"
    ws_info["B1"] = result.get("modelType", "")
    ws_info["A2"] = "Overall Confidence"
    ws_info["B2"] = result.get("confidence", 0)
    ws_info["A3"] = "Status"
    ws_info["B3"] = result.get("status", "")
    ws_info["A4"] = "Pages Processed"
    ws_info["B4"] = len(result.get("pages", []))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
